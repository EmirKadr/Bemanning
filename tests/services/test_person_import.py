import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.backend.routers.persons import build_person_import_template_excel, parse_person_import_excel


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_build_person_import_template_excel_has_expected_headers():
    workbook = load_workbook(io.BytesIO(build_person_import_template_excel()))
    sheet = workbook.active

    assert [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value] == [
        "namn",
        "hemomr\u00e5de",
        "huvudst\u00e4lle",
        "sortering",
    ]


def test_parse_person_import_excel_accepts_name_only():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn", "hemomr\u00e5de", "huvudst\u00e4lle", "sortering"],
                ["Anna Andersson", None, None, None],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].name == "Anna Andersson"
    assert rows[0].home_area is None
    assert rows[0].home_activity is None
    assert rows[0].sort_order is None


def test_parse_person_import_excel_accepts_optional_fields():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn", "hemomr\u00e5de", "huvudst\u00e4lle", "sortering"],
                ["Bo Berg", "GG", "GG VM", 12],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].home_area == "GG"
    assert rows[0].home_activity == "GG VM"
    assert rows[0].sort_order == 12


def test_parse_person_import_excel_collects_row_errors():
    rows, errors = parse_person_import_excel(
        workbook_bytes(
            [
                ["namn", "sortering"],
                [None, 1],
                ["Cecilia", "1,5"],
            ]
        )
    )

    assert rows == []
    assert [error.row for error in errors] == [2, 3]
    assert "Namn" in errors[0].error
    assert "heltal" in errors[1].error


def test_parse_person_import_excel_requires_name_header():
    with pytest.raises(HTTPException) as exc:
        parse_person_import_excel(workbook_bytes([["hemomr\u00e5de"], ["GG"]]))

    assert exc.value.status_code == 400
