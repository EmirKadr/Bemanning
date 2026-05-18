import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.backend.routers.activities import build_activity_import_template_excel, parse_activity_import_excel


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_build_activity_import_template_excel_has_expected_headers():
    workbook = load_workbook(io.BytesIO(build_activity_import_template_excel()))
    sheet = workbook.active

    assert [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value, sheet["E1"].value, sheet["F1"].value] == [
        "etikett",
        "område",
        "summeras som",
        "kategori",
        "färg",
        "sortering",
    ]


def test_parse_activity_import_excel_accepts_label_only():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "område", "summeras som", "kategori", "färg", "sortering"],
                ["GG Påfyllning", None, None, None, None, None],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].label == "GG Påfyllning"
    assert rows[0].area is None
    assert rows[0].summary_activity is None
    assert rows[0].category == "work"
    assert rows[0].color == "#ffffff"
    assert rows[0].sort_order is None


def test_parse_activity_import_excel_accepts_optional_fields():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "område", "summeras som", "kategori", "färg", "sortering"],
                ["Frånvaro", "GG", "Ledigt", "frånvaro", "fee2e2", 20],
            ]
        )
    )

    assert errors == []
    assert len(rows) == 1
    assert rows[0].area == "GG"
    assert rows[0].summary_activity == "Ledigt"
    assert rows[0].category == "absence"
    assert rows[0].color == "#fee2e2"
    assert rows[0].sort_order == 20


def test_parse_activity_import_excel_collects_row_errors():
    rows, errors = parse_activity_import_excel(
        workbook_bytes(
            [
                ["etikett", "kategori", "färg", "sortering"],
                [None, "arbete", "#ffffff", 1],
                ["Okänd kategori", "annat", "#ffffff", 2],
                ["Fel färg", "arbete", "gul", 3],
                ["Fel sort", "arbete", "#ffffff", "1,5"],
            ]
        )
    )

    assert rows == []
    assert [error.row for error in errors] == [2, 3, 4, 5]
    assert "Etikett" in errors[0].error
    assert "Kategori" in errors[1].error
    assert "Färg" in errors[2].error
    assert "heltal" in errors[3].error


def test_parse_activity_import_excel_requires_label_header():
    with pytest.raises(HTTPException) as exc:
        parse_activity_import_excel(workbook_bytes([["område"], ["GG"]]))

    assert exc.value.status_code == 400
